import openpyxl
import os

def GetRowCount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_row

def GetColCount(path,sheetname):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.max_column

def GetCellData(path,sheetname,rownum,colnum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum,column=colnum).value

def WriteCellData(path,sheetname,rownum,colnum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=colnum).value = data
    workbook.save(path)

path1 = "testdata.xlsx"
sheetname1 = "Logintest"

rows = GetRowCount(path1,sheetname1)
cols = GetColCount(path1,sheetname1)

print(rows,"-------",cols)

print(GetCellData(path1,sheetname1,2,2))

WriteCellData(path1,sheetname1,5,5,"Writing to Cell")